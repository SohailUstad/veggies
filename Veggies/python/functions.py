#create account
#login

import os
from openpyxl import Workbook, load_workbook
import random

class Account:
    def __init__(self):
        self.accountNo = ""
        self.fullName = "Unknown"
        self.pincode = ""
        self.balance = 0
        self.file_path = 'data.xlsx'
        self.workbook = load_workbook(self.file_path)
    def createAccount(self):
        self.fullName = input("Enter your fullname: ")
        pincode = input("Create a 4 digit pin: ")
        confirmPin = input("Confirm pin: ")
        if pincode != confirmPin:
            print("Pin doesn't match. Please try again the whole process")
            return
        else:
            self.pincode = pincode
            self.accountNo = random.randint(100000000000, 999999999999)
            acn = self.saveInExcel()
            if acn > 0:
                print("Account Created")
            else:
                print("Something went wrong")

    def saveInExcel(self):
        file_path = 'data.xlsx'
        if not os.path.exists(file_path):
            self.createExcelFile(file_path)
        data = self.prepareDataToStoreInExcel()
        wb = load_workbook(file_path)
        ws = wb.active
        #next_row = ws.max_row + 1
        for d in data:
            ws.append(d)
        wb.save(file_path)
        return self.accountNo

    def prepareDataToStoreInExcel(self):
        new_data = [
            (self.accountNo,
             self.fullName,
             self.pincode,
             self.balance
             )
            ]
        return new_data

    def createExcelFile(self,file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "AccountHoldersData"
        ws['A1'] = 'Account Number'
        ws['B1'] = 'Full Name'
        ws['C1'] = 'Pincode'
        ws['D1'] = 'Current Balance'
        wb.save(file_path)
        print(f"{file_path} created successfully.")

    def login(self,acn,pin):
        sheet = self.getAccountDetailsSheet()
        account_details_row_index = self.checkIfAccountNumberExistsOrNot(acn,sheet)
        if account_details_row_index>=2:
            account_details = []
            for col in range(1, sheet.max_column + 1):  # Iterate through all columns
                cell_value = sheet.cell(row=account_details_row_index, column=col).value
                account_details.append(cell_value)
            if account_details[2] != pin:
                print("Invalid Pin")
                return False
            self.accountNo = account_details[0]
            self.fullName = account_details[1]
            self.balance = account_details[3]
            print("Login Success")
            return True
        else:
            print("Account Number Does not exists")
            return False
        
    
    def printDetailsAfterLogin(self):
        print(self.fullName)
        print(self.accountNo)
        print(self.balance)

    def deposit(self,amount):
        self.balance = self.balance + amount
        self.updateBalance()
        print("Amount Deposited Successfully")

    def withdraw(self,amount):
        if amount>self.balance:
            print("Insufficient balance")
            return
        else:
            self.balance = self.balance - amount
        self.updateBalance()
        print("Amount Withdrawn Successfully")


    def getAccountDetailsSheet(self):
        sheet = self.workbook['AccountHoldersData']
        return sheet

    def checkIfAccountNumberExistsOrNot(self,acn,sheet):
        account_number_column_index = 1
        list_of_account_numbers = []
        for r in range(2, sheet.max_row + 1):  # Iterate through all rows
            cell_value = sheet.cell(row=r, column=account_number_column_index).value
            list_of_account_numbers.append(str(cell_value))
        if str(acn) not in list_of_account_numbers:
            return 0
        account_details_row_index = list_of_account_numbers.index(acn)+2
        return account_details_row_index
    
    def updateBalance(self):
        sheet = self.getAccountDetailsSheet()
        accoun_details_row_index = self.checkIfAccountNumberExistsOrNot(str(self.accountNo),sheet)
        balance_col_index = 4
        print("acn row: ",accoun_details_row_index)
        sheet.cell(row=accoun_details_row_index, column=balance_col_index).value = self.balance
        self.workbook.save(filename=self.file_path)



ac = Account()


while True:
    choice = int(input("Hi, Welcome to my bank\nType\n1. Create New Account\n2. Login\n3. Exit"))
    if choice == 1:
        ac.createAccount()
    elif choice == 2:
        acn = input("Enter account number: ")
        pin = input("Enter Pin")
        isLogin = ac.login(acn,pin)
        while True:
            if isLogin:
                print("Hi!!!")
                ac.printDetailsAfterLogin()
                print("Type\n1. Check Balance\n2. Deposit\n3. Withdraw\n4. See transaction history\n5. Logout")
                innerChoice = int(input())
                if innerChoice == 1:
                    print("Not Available")
                elif innerChoice == 2:
                    amnt = int(input("Enter amount to deposite"))
                    ac.deposit(amnt)
                elif innerChoice == 3:
                    amnt = int(input("Enter amount to withdraw"))
                    ac.withdraw(amnt)
                elif innerChoice == 4:
                    print("Not Available")
                elif innerChoice == 5:
                    print("Bye bye !!!")
                    break
                else:
                    print("Invalid input")
    elif choice == 3:
        break
    else:
        print("Invalid input")

